"""Generate seed-style SQL from platser_och_aktiviteter.xlsx.

Reads the workbook's three kinds of sheets and emits INSERT statements in the
same style as server/priv/seeding/*.sql:

- "Platser"            -> location rows (+ opening_hours JSONB per date)
- "Kategorier"         -> location_tag rows
- one sheet per day    -> activity rows

The `activity` table only has columns for title, description, max_attendees,
start_time and end_time (see server/priv/migrations). The day sheets carry a
lot more per-activity data than that (English text, age-group eligibility,
Kategori, Plats, Kraver bokning, Dold i appen, Agare) but there is nowhere to
put most of it yet, so it is intentionally dropped. The one exception is
Kategori + Plats: since location_tag can already be linked to a location,
each activity's category is recorded as a tag on the location it takes place
at.

Requires: pip install openpyxl

Usage: python3 generate_import.py
Output: import_locations.sql, import_activities.sql (next to this script)
"""

import re
import sys
import uuid
from datetime import date, time
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = SCRIPT_DIR / "platser_och_aktiviteter.xlsx"
LOCATIONS_OUTPUT_PATH = SCRIPT_DIR / "import_locations.sql"
ACTIVITIES_OUTPUT_PATH = SCRIPT_DIR / "import_activities.sql"

YEAR = 2026
MONTH_NAMES = {"juli": 7, "augusti": 8}
DAY_SHEET_NAMES = [
    "25 juli",
    "26 juli",
    "27 juli",
    "28 juli",
    "29 juli",
    "30 juli",
    "31 juli",
    "1 augusti",
]

# Placeholders for location/tag fields the spreadsheet doesn't provide.
DEFAULT_ICON_NAME = "tabler-map-pin"
DEFAULT_ICON_VARIANT = "outline"
DEFAULT_COLOR = "#6b7280"
# Used for locations with no latitude/longitude in the sheet, so they still
# satisfy the NOT NULL constraint. Deliberately an obviously-wrong value
# (middle of the ocean) rather than a real camp coordinate, so unfixed rows
# are easy to spot on a map.
FALLBACK_LATITUDE = 0.0
FALLBACK_LONGITUDE = 0.0

TIME_RANGE_RE = re.compile(r"(\d{1,2})[:.](\d{2})\s*-\s*(\d{1,2})[:.](\d{2})")


def new_id() -> str:
    return str(uuid.uuid7())


def parse_swedish_date(label: str) -> date:
    day_str, month_name = label.strip().split(" ", 1)
    return date(YEAR, MONTH_NAMES[month_name.strip()], int(day_str))


def parse_opening_hours_cell(text: str | None) -> list[tuple[str, str]]:
    """Extract (from, to) HH:MM pairs from a free-text opening-hours cell.

    Cells mix separators (space, comma, newline) and typos (`.` instead of
    `:`, unpadded hours), so this scans for "H(:|.)MM - H(:|.)MM" patterns
    rather than splitting on a fixed delimiter.
    """
    if not text:
        return []
    ranges = []
    for from_h, from_m, to_h, to_m in TIME_RANGE_RE.findall(text):
        ranges.append((f"{int(from_h):02d}:{from_m}", f"{int(to_h):02d}:{to_m}"))
    return ranges


def sql_str(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_int(value: int | None) -> str:
    return "NULL" if value is None else str(value)


def sql_float(value: float) -> str:
    return repr(float(value))


def sql_timestamp(d: date, t: time) -> str:
    return f"'{d.isoformat()} {t.strftime('%H:%M:%S')}'"


def sql_jsonb_opening_hours(opening_hours: dict[str, list[tuple[str, str]]]) -> str:
    if not opening_hours:
        return "'{}'::jsonb"
    date_entries = []
    for iso_date in sorted(opening_hours):
        ranges = ",\n".join(
            f'                {{"from": "{f}", "to": "{t}"}}'
            for f, t in opening_hours[iso_date]
        )
        date_entries.append(f'            "{iso_date}": [\n{ranges}\n            ]')
    body = ",\n".join(date_entries)
    return f"'{{\n{body}\n        }}'::jsonb"


def read_locations(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Platser"]
    rows = list(ws.iter_rows(values_only=True))
    day_headers = rows[1][1:9]
    locations = []
    for row in rows[2:]:
        name = row[0]
        if not name:
            continue
        name = name.strip()
        opening_hours = {}
        for header, cell in zip(day_headers, row[1:9]):
            ranges = parse_opening_hours_cell(cell)
            if ranges:
                opening_hours[parse_swedish_date(header).isoformat()] = ranges
        description = (row[9] or "").strip()
        latitude = row[10]
        longitude = row[11]
        locations.append(
            {
                "id": new_id(),
                "name": name,
                "description": description,
                "latitude": float(latitude) if latitude else None,
                "longitude": float(longitude) if longitude else None,
                "opening_hours": opening_hours,
            }
        )
    return locations


def read_categories(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Kategorier"]
    rows = list(ws.iter_rows(values_only=True))
    names = [row[0].strip() for row in rows[2:] if row[0]]
    return [{"id": new_id(), "name": name} for name in names]


def read_activities(wb: openpyxl.Workbook) -> tuple[list[dict], list[str]]:
    activities = []
    skipped = []
    for sheet_name in DAY_SHEET_NAMES:
        ws = wb[sheet_name]
        activity_date = parse_swedish_date(sheet_name)
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            name = row[0]
            if not name or str(name).startswith("Exempel"):
                continue
            name = name.strip()
            start_time, end_time = row[12], row[13]
            if start_time is None or end_time is None:
                skipped.append(f"{sheet_name}: {name} (missing start/end time)")
                continue
            max_attendees = row[14] if isinstance(row[14], int) else None
            activities.append(
                {
                    "id": new_id(),
                    "title": name,
                    "description": (row[2] or "").strip(),
                    "max_attendees": max_attendees,
                    "start": (activity_date, start_time),
                    "end": (activity_date, end_time),
                    "category": row[4].strip() if row[4] else None,
                    "location_name": row[11].strip() if row[11] else None,
                }
            )
    return activities, skipped


def validate_references(
    locations_by_name: dict[str, dict],
    category_names: set[str],
    activities: list[dict],
) -> None:
    """Fail fast if an activity's Plats/Kategori isn't in the Platser/
    Kategorier sheets, rather than silently inventing a placeholder row."""
    missing_locations = sorted(
        {
            activity["location_name"]
            for activity in activities
            if activity["location_name"]
            and activity["location_name"] not in locations_by_name
        }
    )
    missing_categories = sorted(
        {
            activity["category"]
            for activity in activities
            if activity["category"] and activity["category"] not in category_names
        }
    )
    if missing_locations:
        print(
            "Error: activities reference Plats not found in the \"Platser\" sheet:",
            file=sys.stderr,
        )
        for name in missing_locations:
            print(f"  - {name}", file=sys.stderr)
    if missing_categories:
        print(
            "Error: activities reference Kategori not found in the \"Kategorier\" sheet:",
            file=sys.stderr,
        )
        for name in missing_categories:
            print(f"  - {name}", file=sys.stderr)
    if missing_locations or missing_categories:
        sys.exit(1)


def render_locations_sql(
    locations_by_name: dict[str, dict],
    categories: list[dict],
    tag_links: set[tuple[str, str]],
    missing_coords: list[str],
) -> str:
    lines = [
        "-- Generated by scripts/generate_import.py from platser_och_aktiviteter.xlsx.",
        "-- Do not edit by hand; re-run the generator instead.",
    ]
    if missing_coords:
        lines.append(
            "-- Fell back to the camp-center coordinates below (no latitude/"
            "longitude in the sheet), fix up manually: "
            + ", ".join(sorted(missing_coords))
        )
    lines.append(
        "-- name_en/description_en/icon_name/color/icon_variant have no "
        "source column in the sheet; placeholders are used below and need "
        "manual review."
    )
    lines.append("")

    locations = list(locations_by_name.values())
    location_values = []
    for loc in locations:
        lat = loc["latitude"] if loc["latitude"] is not None else FALLBACK_LATITUDE
        lon = loc["longitude"] if loc["longitude"] is not None else FALLBACK_LONGITUDE
        location_values.append(
            "    (\n"
            f"        {sql_str(loc['id'])},\n"
            f"        {sql_str(loc['name'])},\n"
            f"        {sql_str(loc['name'])},\n"
            f"        {sql_str(loc['description'])},\n"
            f"        {sql_str(loc['description'])},\n"
            f"        {sql_str(DEFAULT_ICON_NAME)},\n"
            f"        {sql_str(DEFAULT_ICON_VARIANT)},\n"
            f"        {sql_str(DEFAULT_COLOR)},\n"
            f"        {sql_float(lat)},\n"
            f"        {sql_float(lon)},\n"
            f"        {sql_jsonb_opening_hours(loc['opening_hours'])}\n"
            "    )"
        )
    lines.append(
        "INSERT INTO location (\n"
        "        id,\n"
        "        name,\n"
        "        name_en,\n"
        "        description,\n"
        "        description_en,\n"
        "        icon_name,\n"
        "        icon_variant,\n"
        "        color,\n"
        "        latitude,\n"
        "        longitude,\n"
        "        opening_hours\n"
        "    )\n"
        "VALUES\n" + ",\n".join(location_values) + ";"
    )
    lines.append("")

    category_values = [
        "    (\n"
        f"        {sql_str(cat['id'])},\n"
        f"        {sql_str(cat['name'])},\n"
        f"        {sql_str(cat['name'])},\n"
        f"        {sql_str(DEFAULT_ICON_NAME)},\n"
        f"        {sql_str(DEFAULT_ICON_VARIANT)}\n"
        "    )"
        for cat in categories
    ]
    lines.append(
        "INSERT INTO location_tag (id, name, name_en, icon_name, icon_variant)\n"
        "VALUES\n" + ",\n".join(category_values) + ";"
    )
    lines.append("")

    if tag_links:
        category_by_name = {cat["name"]: cat for cat in categories}
        link_values = [
            "    (\n"
            f"        {sql_str(category_by_name[cat_name]['id'])},\n"
            f"        {sql_str(locations_by_name[loc_name]['id'])}\n"
            "    )"
            for cat_name, loc_name in sorted(tag_links)
        ]
        lines.append(
            "-- Derived from the Kategori column of activities held at each location;\n"
            "-- location_tag has no other link to activities today.\n"
            "INSERT INTO location_tag_location (location_tag_id, location_id)\n"
            "VALUES\n" + ",\n".join(link_values) + ";"
        )
        lines.append("")

    return "\n".join(lines)


def render_activities_sql(activities: list[dict], skipped: list[str]) -> str:
    lines = [
        "-- Generated by scripts/generate_import.py from platser_och_aktiviteter.xlsx.",
        "-- Do not edit by hand; re-run the generator instead.",
        "--",
        "-- Only title, description, max_attendees, start_time and end_time are",
        "-- imported: the activity table has no columns yet for English text,",
        "-- age-group eligibility, Kategori, Plats, Kraver bokning, Dold i appen or",
        "-- Agare, so those columns from the sheet are dropped. Kategori/Plats are",
        "-- captured indirectly as location tags -- see import_locations.sql.",
    ]
    if skipped:
        lines.append("--")
        lines.append("-- Skipped (missing start/end time in the sheet):")
        for entry in skipped:
            lines.append(f"--   {entry}")
    lines.append("")

    values = []
    for activity in activities:
        start_date, start_time = activity["start"]
        end_date, end_time = activity["end"]
        values.append(
            "    (\n"
            f"        {sql_str(activity['id'])},\n"
            f"        {sql_str(activity['title'])},\n"
            f"        {sql_str(activity['description'])},\n"
            f"        {sql_int(activity['max_attendees'])},\n"
            f"        {sql_timestamp(start_date, start_time)},\n"
            f"        {sql_timestamp(end_date, end_time)}\n"
            "    )"
        )
    lines.append(
        "INSERT INTO activity (\n"
        "        id,\n"
        "        title,\n"
        "        description,\n"
        "        max_attendees,\n"
        "        start_time,\n"
        "        end_time\n"
        "    )\n"
        "VALUES\n" + ",\n".join(values) + ";"
    )
    lines.append("")

    return "\n".join(lines)


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

    locations = read_locations(wb)
    categories = read_categories(wb)
    activities, skipped = read_activities(wb)

    locations_by_name = {loc["name"]: loc for loc in locations}
    category_names = {cat["name"] for cat in categories}
    validate_references(locations_by_name, category_names, activities)

    missing_coords = sorted(
        name
        for name, loc in locations_by_name.items()
        if loc["latitude"] is None or loc["longitude"] is None
    )

    tag_links = {
        (activity["category"], activity["location_name"])
        for activity in activities
        if activity["category"] and activity["location_name"]
    }

    LOCATIONS_OUTPUT_PATH.write_text(
        render_locations_sql(locations_by_name, categories, tag_links, missing_coords)
    )
    ACTIVITIES_OUTPUT_PATH.write_text(render_activities_sql(activities, skipped))

    print(f"Wrote {len(locations_by_name)} locations -> {LOCATIONS_OUTPUT_PATH}")
    print(f"Wrote {len(categories)} tags -> {LOCATIONS_OUTPUT_PATH}")
    print(f"Wrote {len(tag_links)} location<->tag links -> {LOCATIONS_OUTPUT_PATH}")
    print(f"Wrote {len(activities)} activities -> {ACTIVITIES_OUTPUT_PATH}")
    if skipped:
        print(f"Skipped {len(skipped)} activities missing start/end time:")
        for entry in skipped:
            print(f"  - {entry}")


if __name__ == "__main__":
    main()
